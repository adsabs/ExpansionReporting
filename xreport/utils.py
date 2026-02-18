import re
import os
import sys
import time
import urllib.request, urllib.parse, urllib.error
import requests
import math
import functools
from datetime import date
from adsgcon.gmanager import GoogleManager
import openpyxl
from openpyxl.styles import Font
from urllib.parse import urlencode, urlunparse
# ============================= INITIALIZATION ==================================== #

from adsputils import setup_logging, load_config

proj_home = os.path.realpath(os.path.join(os.path.dirname(__file__), '../'))
config = load_config(proj_home=proj_home)
logger = setup_logging(__name__, proj_home=proj_home,
                        level=config.get('LOGGING_LEVEL', 'INFO'),
                        attach_stdout=config.get('LOG_STDOUT', False))
# Exception definitions
class GoogleUploadException(Exception):
    pass

class GoogleManagerException(Exception):
    pass

class FolderIdNotFound(Exception):
    pass
# =============================== HELPER FUNCTIONS ================================ #
def retry(tries=4, delay=3, backoff=1):
    """A so-called decorator function implementing a retry-on-exception functionality"""
    # tries: the number of attempts to retry
    # delay: the number of seconds to wait between retries
    # backoff: a multiplier option (larger than 1 means that wait times increase by this factor)
    # logger: file handle of log file (None means that messages are written to stdout)
    def deco_retry(func):
        @functools.wraps(func)
        def f_retry(*args, **kwargs):
            m_tries, m_delay = tries, delay

            while m_tries > 1:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    msg = f"Exception '{e}' while querying API, retrying in {m_delay} seconds..."
                    logger.warning(msg)
                    time.sleep(m_delay)
                    m_tries -= 1
                    m_delay *= backoff

            return func(*args, **kwargs)

        return f_retry
    return deco_retry

def _group(lst, n):
    """
    Transform a list of values into a list of tuples of length n
    
    param: lst: the input list
    param: n: tuple length
    """
    for i in range(0, len(lst), n):
        val = lst[i:i+n]
        if len(val) == n:
            yield tuple(val)

def _string2list(numstr):
    """
    Convert a string of numbers into a list/range
    Example: '1,2,5-7,10' --> [1, 2, 5, 6, 7, 10]
    
    param: str: the input string
    """
    result = []
    for part in numstr.split(','):
        if '-' in part:
            a, b = part.split('-')
            a, b = int(a), int(b)
            result.extend(range(a, b + 1))
        else:
            a = int(part)
            result.append(a)
    return result

def _make_dict(tup, key_is_int=True):
    """
    Turn list of tuples into a dictionary
    
    param: tup: list of tuples
    param: key_is_int: keys of dictionary will be integers if true
    """
    newtup = tup
    if key_is_int:
        newtup = [(int(re.sub("[^0-9]", "", e[0])), e[1]) for e in tup]        
    return dict(newtup)

@retry(tries=5)
def _do_query(conf, params, endpoint='search/query'):
    """
    Send of a query to the ADS API (essentially, any API defined by config values)
    
    param: conf: dictionary with configuration values
    param: params: idctionary with query parameters
    """
    headers = {}
    headers["Authorization"] = "Bearer {}".format(conf['ADS_API_TOKEN'])
    headers["Accept"] = "application/json"
    if isinstance(params, str):
        url = "{}/{}/{}".format(conf['ADS_API_URL'], endpoint, params)
    else:
        url = "{}/{}?{}".format(conf['ADS_API_URL'], endpoint, urllib.parse.urlencode(params))
    r_json = {}
    try:
        r = requests.get(url, headers=headers)
    except Exception as err:
        logger.error("Search API request failed: {}: {}".format(err, url))
        raise
    if not r.ok:
        msg = "Search API request with error code '{}': {}".format(r.status_code, url)
        logger.error(msg)
        raise Exception(msg)
    else:
        try:
            r_json = r.json()
        except ValueError:
            msg = "No JSON object could be decoded from Search API"
            logger.error(msg)
            raise Exception(msg)
        else:
            return r_json
    return r_json

# =============================== DATA RETRIEVAL FUNCTIONS ==================== #

def _get_citations(conf, query_string):
    """
    Get citation counts via ADS API pivot query
    
    param: conf: dictionary with configuration values
    param: query_string: the query string to execute pivot query on
    """
    # creating the date object of today's date
    todays_date = date.today()
    params = {
            'facet': 'true',
            'facet.limit': 2000,
            'facet.minCount': '1',
            'facet.pivot': 'year,citation_count',
            'q': query_string,
            'sort': 'citation_count desc'
        }
    query_data = _do_query(conf, params)
    pivots_data = query_data['facet_counts']['facet_pivot'].get('year,citation_count')
    data =[item for sublist in [p['pivot'] for p in pivots_data] for item in sublist]
    # The total number of citations is the sum of each citation value times its multiplicity
    total_cites = sum([e['count']*e['value'] for e in data])
    
    return total_cites

def _get_facet_data(conf, query_string, facet):
    """
    Do an ADS API facet query
    
    param: conf: dictionary with configuration values
    param: query_string: the query string to execute pivot query on
    param: facet: the facet to return
    """
    params = {
        'q':query_string,
        'fl': 'id',
        'rows': 1,
        'facet':'on',
        'facet.field': facet,
        'facet.limit': 1000,
        'facet.mincount': 1,
        'facet.offset':0,
        'sort':'date desc'
    }

    data = _do_query(conf, params)
    results = data['facet_counts']['facet_fields'].get(facet)
    # Return a dictionary with facet values and associated frequencies
    res_dict = _make_dict(list(_group(results, 2)))
    if facet == 'volume':
        try:
            filt_dict = {key:value for (key, value) in res_dict.items() if key < 2100}
        except:
            filt_dict = res_dict
        return filt_dict
    else:
        return res_dict

def _get_records(conf, query_string, return_fields):
    """
    Do a general ADS API query
    
    param: conf: dictionary with configuration values
    param: query_string: the query string to execute pivot query on
    param: return_fields: which Solr fields to return
    """
    start = 0
    rows = 1000
    results = []
    params = {
        'q':query_string,
        'fl': return_fields,
        'rows': rows,
        'start': start
    }
    data = _do_query(conf, params)
    try:
        results = data['response']['docs']
    except:
        raise Exception('Solr returned unexpected data!')
    num_documents = int(data['response']['numFound'])
    num_paginates = int(math.ceil((num_documents) / (1.0*rows))) - 1
    start += rows
    for i in range(num_paginates):
        params['start'] = start
        data = _do_query(conf, params)
        try:
            results += data['response']['docs']
        except:
            raise Exception('Solr returned unexpected data!')
        start += rows
    return results

def _get_usage(config, jrnls=[], bibcodes=[], udata='reads'):
    """
    Return usage data from Classic index files for a set of journals of bibcodes
    
    param: conf: dictionary with configuration values
    param: journals: a list of bibstems, if specified
    param: bibcodes: a list of bibcodes, if specified
    param: udata: what type of usage data to return
    """
    total = 0
    recent = 0
    index_file = config.get('CLASSIC_USAGE_INDEX')[udata]
    # Cycle through index file and get usage data for either specific journals or bibcodes
    with open(index_file) as fh:
        for line in fh:
           data = line.strip().split('\t')
           if jrnls and data[0][4:9] not in jrnls:
               continue
           elif bibcodes and data[0] not in bibcodes:
               continue
           total += sum([int(d) for d in data[1:]])
           recent += int(data[-1])
    return total, recent

def _get_journal_coverage(conf, jrnl):
    """
    Get metadata completeness statistics from Journals Database for a given journal
    
    param: conf: dictionary with configuration values
    param: jrnl: a journal abbreviation (bibstem)
    """
    try:
        data = _do_query(conf, jrnl.replace('.',''), endpoint='journals/summary')
        completeness_data = data['summary']['master']['completeness_details']['completeness_by_volume']
    except:
        logger.error('No completeness data found for: {0}'.format(jrnl))
        completeness_data = {}

    return completeness_data
    
def _upload_to_teamdrive(coll,subj,excel_file):
    """
    Upload Excel report to Team Drive and do some cleanup
    
    param: coll: Collection name
    param: subj: Data type (fulltext, metadata or references)
    param: excel_file: Excel report with full path
    """
    logger.info('Uploading {0} to Team Drive'.format(excel_file))
    name2id = config['NAME2ID']
    try:
        gm = GoogleManager(authtype="service",
                       folderId=config['FOLDER_ID'],
                       secretsFile=config['SECRETS_FILE'],
                       scopes=config['SCOPES'])
    except Exception as err:
        logger.error('Failed to instantiate GoogleManager: {0}'.format(err))
        raise GoogleManagerException('Failed to instantiate GoogleManager: {0}'.format(err))
    try:
        folderIdent = name2id[coll]
    except Exception as err:
        raise FolderIdNotFound("Cannot find folder id for: {0}".format(topic))
    # Point the GoogleManager object to this folder
    gm.folderid = folderIdent
    # Get all reports stored in this folder (with their Google Drive ID and name)
    reports = gm.list_files(return_fields="id,name")
    # Get all reports in the Google Drive for the data type at hand
    reports = [r for r in reports if r['name'].lower().startswith(subj)]
    # Is the newly generated report already in the Team Drive?
    if os.path.basename(excel_file).replace('.xlsx','') in [r['name'] for r in reports]:
        # Report is already in Team Drive. Nothing to do.
        return False
    # This report is not yet on the Team Drive. Time to upload.
    upload_name = os.path.basename(excel_file)
    kwargs = {
        "infile": excel_file,
        "upload_name": upload_name,
        "mtype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "meta_mtype": "application/vnd.google-apps.spreadsheet"
    }
    try:
        res = gm.upload_file(**kwargs)
        logger.info('Successfully uploaded {0} to Team Drive'.format(excel_file))
    except Exception as err:
        raise GoogleUploadException("Unable to upload file %s to google drive: %s" % (excel_file if excel_file else '?', err)) from None
    # Finally, remove the reports that were there already
    for r in reports:
        try:
            res = gm.delete_file(fileId=r['id'])
            logger.info('Successfully deleted {0} from Team Drive'.format(r['name']))
        except Exception as err:
            raise GoogleDeleteException("Unable to delete file %s from google drive: %s" % (r['name'], err))
    return True

def _create_url(bibstem, year_vol):
    """
    Create SciX URL for records without fulltext for a given year or volume
    
    param: bibstem: The bibstem for the journal in question
    param: year_vol: Filter by year or volume
    """
    scheme = 'https'
    netloc = 'www.scixplorer.org'
    path = '/search'
    params = {
        'p': '1',
        'q': 'bibstem:{0} {1} -has:body -title:(erratum OR editorial) doctype:article'.format(bibstem, year_vol),
        'sort': ['score desc', 'date desc'], # List handles multiple sort params
        'd': 'general'
    }
    query_string = urlencode(params, doseq=True)
    final_url = urlunparse((scheme, netloc, path, '', query_string, ''))
    return final_url

def _add_hyperlinks(filename, sheet_name='Sheet1', threshold=80):
    """
    Add hyperlinks to coverage spreadsheet for cells with values smaller than a threshold
    
    """
    # Skip these rows
    skip = ['publisher ->','start year ->','last year ->','start vol ->','last vol ->']
    # 1. Load the workbook and select the active worksheet
    wb = openpyxl.load_workbook(filename)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        logger.info('Sheet {0} not found. Using active sheet.'.format(sheet_name))
        ws = wb.active

    # 2. Extract column headers (assuming the first row is the header row)
    # Use ws.iter_rows(min_row=1, max_row=1, values_only=True) for efficiency
    column_headers = {}
    for cell in ws[1]:
        column_headers[cell.column_letter] = cell.value

    # 3. Extract row headers (assuming the first column is the header column)
    row_headers = {}
    for row_cell in ws.iter_rows(min_col=1, max_col=1, min_row=2): # Start from the second row
        cell = row_cell[0]
        row_headers[cell.row] = cell.value

    # 4. Iterate over all data cells (excluding the header row and column)
    # Start from the second row (min_row=2) and second column (min_col=2)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            try:
                # Try to convert cell value to a number for the condition check
                cell_value = float(cell.value)
                # Check if the value is between 0 and threshold (exclusive of 0, inclusive of 60 based on prompt)
                if 0 < cell_value <= float(threshold):
                    # Get the corresponding column and row headers
                    col_header = column_headers.get(cell.column_letter, 'N/A')
                    row_header = row_headers.get(cell.row, 'N/A')
                    if row_header not in skip:
                        # Determine whether we need to filter by year or volume (from file name)
                        if 'year' in filename:
                            filter = "year:{0}".format(row_header)
                        else:
                            filter = "volume:{0}".format(row_header)
                        # Assign hyperlink to cell
                        cell.hyperlink = _create_url(col_header, filter) 
                        # Apply the standard hyperlink style (blue and underlined)
                        cell.font = Font(color="0000FF", underline="single")
            except (ValueError, TypeError):
                # Handle cases where the cell value is not a number
                continue
    # 5. Save the modified workbook
    output_filename = filename.replace('.xlsx','.modified.xlsx')
    wb.save(output_filename)
    # 6. Replace the original workbook with the hyperlinked one
    try:
        # Renames source to destination, replacing destination if it exists
        os.replace(output_filename, filename)
    except OSError as err:
        logger.error('Error replacing {0} by {1}: {2}'.format(filename, output_filename, err))